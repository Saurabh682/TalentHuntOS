import csv
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.actions.api import dispatch_action
from app.actions.models import ActionExecution
from app.analytics.artifacts import get_report_download, resolve_report_path
from app.analytics.models import ReportArtifact
from app.analytics.routes import download_report_artifact
from app.candidates.models import Candidate
from app.hunts.models import HuntCandidate, HuntStage, TalentHunt
from app.infrastructure.db import Base


def _factory(tmp_path):
    engine = create_engine(f"sqlite:///{tmp_path / 'reports.db'}")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)


def _use_factory(monkeypatch, factory, tmp_path):
    monkeypatch.setattr("app.infrastructure.db.SessionFactory", factory)
    monkeypatch.setattr("app.actions.locks.SessionFactory", factory)
    monkeypatch.setattr("app.actions.approvals.SessionFactory", factory)
    monkeypatch.setattr("app.actions.tool_calls.SessionFactory", factory)
    monkeypatch.setattr("app.analytics.artifacts.REPORTS_DIR", tmp_path / "private-reports")


def _seed(factory):
    with factory() as db:
        hunt = TalentHunt(
            title="=CMD|' /C calc'!A0",
            target_role="Animator",
            status="Active",
        )
        db.add(hunt)
        db.flush()
        candidate = Candidate(full_name="Report Candidate", status="Sourced")
        db.add(candidate)
        db.flush()
        stage = HuntStage(hunt_id=hunt.id, name="Sourced", position=0)
        db.add(stage)
        db.flush()
        db.add(
            HuntCandidate(
                hunt_id=hunt.id,
                stage_id=stage.id,
                candidate_id=candidate.id,
                full_name="Report Candidate",
                source_platform="linkedin",
                match_score=82,
            )
        )
        db.commit()
        return hunt.id


def _create(report_format, hunt_id):
    return dispatch_action(
        "reports.analytics.create",
        {"format": report_format, "hunt_id": hunt_id, "days": 30},
        actor_type="agent",
        session_id="report-tests",
        user_id=1,
    )


def test_report_actions_create_real_safe_artifacts_from_canonical_data(monkeypatch, tmp_path):
    factory = _factory(tmp_path)
    _use_factory(monkeypatch, factory, tmp_path)
    hunt_id = _seed(factory)

    artifacts = {}
    for report_format in ("csv", "xlsx", "pdf"):
        result = _create(report_format, hunt_id)
        assert result.success, result.error
        artifact = result.data["artifact"]
        artifacts[report_format] = artifact
        assert artifact["format"] == report_format
        assert artifact["hunt_id"] == hunt_id
        assert artifact["download_url"] == f"/api/reports/{artifact['id']}"
        assert "relative_path" not in artifact
        assert artifact["provenance"]["source_of_truth"] == (
            "canonical TalentHunt database records"
        )

    csv_path, _ = get_report_download(artifacts["csv"]["id"])
    csv_text = csv_path.read_text(encoding="utf-8-sig")
    rows = list(csv.reader(io.StringIO(csv_text)))
    assert ["Total Candidates Sourced", "1"] in rows
    assert any(row and row[0].startswith("'=CMD") for row in rows)

    xlsx_path, _ = get_report_download(artifacts["xlsx"]["id"])
    assert xlsx_path.read_bytes().startswith(b"PK")
    workbook = load_workbook(xlsx_path, data_only=False)
    assert workbook.sheetnames == [
        "Executive Summary",
        "Funnel",
        "Hunt Velocity",
        "Source Quality",
        "Outreach",
        "Trends",
    ]
    assert workbook["Executive Summary"]["B3"].value.startswith("'=CMD")
    assert workbook["Hunt Velocity"]["A2"].data_type != "f"
    assert workbook["Hunt Velocity"]["D2"].value == 1
    assert workbook["Source Quality"]["A2"].value == "Linkedin"
    assert workbook["Source Quality"]["B2"].value == 1

    pdf_path, _ = get_report_download(artifacts["pdf"]["id"])
    assert pdf_path.read_bytes().startswith(b"%PDF")
    pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(pdf_path).pages)
    assert "TalentHunt OS - Analytics and Intelligence" in pdf_text
    assert "Candidates Sourced" in pdf_text

    with factory() as db:
        assert db.query(ReportArtifact).count() == 3
        assert (
            db.query(ActionExecution)
            .filter_by(action_name="reports.analytics.create", status="completed")
            .count()
            == 3
        )


def test_report_list_get_and_download_integrity_are_bounded(monkeypatch, tmp_path):
    factory = _factory(tmp_path)
    _use_factory(monkeypatch, factory, tmp_path)
    hunt_id = _seed(factory)
    created = _create("csv", hunt_id)
    artifact = created.data["artifact"]

    listed = dispatch_action(
        "reports.list",
        {"format": "csv", "hunt_id": hunt_id, "limit": 5},
        actor_type="agent",
        session_id="report-tests",
        user_id=1,
    )
    fetched = dispatch_action(
        "reports.get",
        {"artifact_id": artifact["id"]},
        actor_type="agent",
        session_id="report-tests",
        user_id=1,
    )
    assert listed.success and listed.data["count"] == 1
    assert listed.data["artifacts"][0]["available"] is True
    assert fetched.success and fetched.data["artifact"]["available"] is True

    path, _ = get_report_download(artifact["id"])
    response = download_report_artifact(artifact["id"])
    assert Path(response.path).resolve() == path.resolve()
    assert response.headers["cache-control"] == "private, no-store"
    content = path.read_bytes()
    path.write_bytes(content[:-1] + bytes([content[-1] ^ 1]))
    with pytest.raises(RuntimeError, match="checksum"):
        get_report_download(artifact["id"])

    with pytest.raises(ValueError, match="escapes"):
        resolve_report_path("../outside.csv")
    assert not (tmp_path / "outside.csv").exists()


def test_report_actions_reject_unknown_scope_and_untrusted_paths(monkeypatch, tmp_path):
    factory = _factory(tmp_path)
    _use_factory(monkeypatch, factory, tmp_path)

    missing = dispatch_action(
        "reports.analytics.create",
        {"format": "pdf", "hunt_id": 999, "days": 30},
        actor_type="agent",
        session_id="report-tests",
        user_id=1,
    )
    invalid = dispatch_action(
        "reports.analytics.create",
        {"format": "html", "output_path": str(tmp_path / "outside.html")},
        actor_type="agent",
        session_id="report-tests",
        user_id=1,
    )
    assert not missing.success and missing.error == "Talent Hunt not found."
    assert not invalid.success
    assert "format" in (invalid.error or "").lower()
    assert not (tmp_path / "outside.html").exists()


def test_report_artifact_path_never_accepts_absolute_paths(tmp_path):
    with pytest.raises(ValueError, match="invalid"):
        resolve_report_path(str(Path(tmp_path / "absolute.csv").resolve()))


def test_analytics_ui_uses_shared_report_action_instead_of_direct_renderers():
    source = Path("app/ui/pages/analytics.py").read_text(encoding="utf-8")
    assert '"reports.analytics.create"' in source
    assert "generate_analytics_pdf" not in source
    assert "generate_analytics_csv" not in source

"""Render canonical TalentHunt analytics as CSV, XLSX, or PDF bytes."""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any
from xml.sax.saxutils import escape


def _generated_at(value: datetime | None) -> datetime:
    current = value or datetime.now(timezone.utc)
    if current.tzinfo is None:
        current = current.replace(tzinfo=timezone.utc)
    return current.astimezone(timezone.utc)


def _display(value: Any) -> str:
    return "" if value is None else str(value)


def _spreadsheet_text(value: Any) -> Any:
    """Keep untrusted text literal in CSV and XLSX consumers."""
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    if stripped.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def generate_analytics_csv(
    metrics_data: dict[str, Any],
    *,
    generated_at: datetime | None = None,
    scope_label: str = "All Talent Hunts",
) -> str:
    """Generate a spreadsheet-safe CSV containing the canonical analytics snapshot."""
    timestamp = _generated_at(generated_at)
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")

    def row(values: list[Any]) -> None:
        writer.writerow([_spreadsheet_text(value) for value in values])

    row(["TalentHunt OS - Executive Analytics Report"])
    row(["Generated At", timestamp.strftime("%Y-%m-%d %H:%M:%S UTC")])
    row(["Scope", scope_label])
    row([])

    kpi = metrics_data.get("kpi", {})
    row(["EXECUTIVE KPI SUMMARY"])
    row(["Metric", "Value"])
    kpi_rows = [
        ("Total Talent Hunts", kpi.get("total_hunts", 0)),
        ("Active Hunts", kpi.get("active_hunts", 0)),
        ("Completed Hunts", kpi.get("completed_hunts", 0)),
        ("Total Candidates Sourced", kpi.get("total_sourced", 0)),
        ("Candidates Interviewing", kpi.get("interviewing_candidates", 0)),
        ("Candidates Hired", kpi.get("hired_candidates", 0)),
        ("Funnel Conversion Rate (%)", kpi.get("conversion_rate", 0)),
        ("Average Time-to-Fill (Days)", kpi.get("avg_time_to_fill_days", 0)),
        ("Outreach Messages Sent", kpi.get("outreach_sent", 0)),
        ("Outreach Replies Received", kpi.get("outreach_replied", 0)),
        ("Outreach Response Rate (%)", kpi.get("response_rate", 0)),
        ("Recorded AI Operations", kpi.get("ai_actions", 0)),
        ("Recorded Estimated Cost Saved (USD)", kpi.get("estimated_cost_saved", 0.0)),
    ]
    for item in kpi_rows:
        row(list(item))
    row([])

    row(["TALENT RECRUITMENT FUNNEL"])
    row(["Stage Name", "Candidate Count", "Overall Conversion (%)", "Drop-off Rate (%)"])
    for stage in metrics_data.get("funnel", {}).get("stages", []):
        row(
            [
                stage.get("stage"),
                stage.get("count"),
                stage.get("overall_conversion"),
                stage.get("dropoff_rate"),
            ]
        )
    row([])

    row(["HUNT VELOCITY AND TIME-TO-FILL"])
    row(
        [
            "Hunt Title",
            "Target Role",
            "Status",
            "Total Candidates",
            "Hired Count",
            "Days Open",
            "Time to Fill (Days)",
        ]
    )
    for hunt in metrics_data.get("velocity", {}).get("hunts_velocity", []):
        row(
            [
                hunt.get("title"),
                hunt.get("target_role"),
                hunt.get("status"),
                hunt.get("total_candidates"),
                hunt.get("hired_count"),
                hunt.get("days_open"),
                hunt.get("time_to_fill_days"),
            ]
        )
    row([])

    row(["SOURCING QUALITY"])
    row(["Source", "Candidate Count"])
    sourcing = metrics_data.get("sourcing", {})
    for source, count in sourcing.get("channels", {}).items():
        row([source, count])
    row([])

    row(["OUTREACH BY CHANNEL"])
    row(["Channel", "Message Count"])
    outreach = metrics_data.get("outreach", {})
    for label, count in outreach.get("channel_counts", {}).items():
        row([label, count])
    row([])
    row(["OUTREACH BY DIRECTION"])
    row(["Direction", "Message Count"])
    for label, count in outreach.get("direction_counts", {}).items():
        row([label, count])
    row([])

    row(["AI ENGINE TELEMETRY"])
    ai_cost = metrics_data.get("ai_cost", {})
    row(["Metric", "Value"])
    for label, key in (
        ("Recorded AI Operations", "total_operations"),
        ("Recorded Local Operations", "local_operations"),
        ("Recorded Cloud Operations", "cloud_operations"),
        ("Recorded Cloud Cost (USD)", "actual_cloud_cost"),
        ("Recorded Cost Saved (USD)", "total_cost_saved"),
    ):
        row([label, ai_cost.get(key, 0)])
    return output.getvalue()


def generate_analytics_excel(
    metrics_data: dict[str, Any],
    *,
    generated_at: datetime | None = None,
    scope_label: str = "All Talent Hunts",
) -> bytes:
    """Generate a real styled Office Open XML workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - guarded by the runtime dependency
        raise RuntimeError("openpyxl is required for XLSX report generation") from exc

    timestamp = _generated_at(generated_at)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Executive Summary"

    navy = "0F172A"
    teal = "0F9D94"
    pale_teal = "DDF7F3"
    pale_blue = "E8F1F8"
    slate = "334155"
    light_border = Side(style="thin", color="CBD5E1")

    def safe(value: Any) -> Any:
        return _spreadsheet_text(value)

    def style_sheet(sheet, widths: list[float]) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 24
        for row_cells in sheet.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=light_border)

    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = "TalentHunt OS - Analytics and Intelligence"
    summary["A1"].fill = PatternFill("solid", fgColor=navy)
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 34
    summary["A2"] = "Generated At"
    summary["B2"] = timestamp.replace(tzinfo=None)
    summary["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary["A3"] = "Scope"
    summary["B3"] = safe(scope_label)
    summary["A5"] = "Metric"
    summary["B5"] = "Value"
    for cell in summary[5]:
        cell.fill = PatternFill("solid", fgColor=teal)
        cell.font = Font(color="FFFFFF", bold=True)

    kpi = metrics_data.get("kpi", {})
    kpi_rows = [
        ("Total Talent Hunts", kpi.get("total_hunts", 0)),
        ("Active Hunts", kpi.get("active_hunts", 0)),
        ("Completed Hunts", kpi.get("completed_hunts", 0)),
        ("Total Candidates Sourced", kpi.get("total_sourced", 0)),
        ("Candidates Interviewing", kpi.get("interviewing_candidates", 0)),
        ("Candidates Hired", kpi.get("hired_candidates", 0)),
        ("Funnel Conversion Rate", (kpi.get("conversion_rate", 0) or 0) / 100),
        ("Average Time-to-Fill", kpi.get("avg_time_to_fill_days", 0)),
        ("Outreach Response Rate", (kpi.get("response_rate", 0) or 0) / 100),
        ("Recorded AI Operations", kpi.get("ai_actions", 0)),
    ]
    for item in kpi_rows:
        summary.append([safe(item[0]), item[1]])
    summary["B12"].number_format = "0.0%"
    summary["B14"].number_format = "0.0%"
    for row_cells in summary.iter_rows(min_row=6, max_row=15, min_col=1, max_col=2):
        for cell in row_cells:
            cell.border = Border(bottom=light_border)
            cell.alignment = Alignment(vertical="center")
    for coordinate in ("B6", "B7", "B8", "B9", "B10", "B11", "B13", "B15"):
        summary[coordinate].number_format = "#,##0"
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 24
    summary.column_dimensions["C"].width = 3
    summary.column_dimensions["D"].width = 3
    summary.freeze_panes = "A6"

    funnel_sheet = workbook.create_sheet("Funnel")
    funnel_sheet.append(
        ["Pipeline Stage", "Candidate Count", "Overall Conversion", "Drop-off Rate"]
    )
    for stage in metrics_data.get("funnel", {}).get("stages", []):
        funnel_sheet.append(
            [
                safe(_display(stage.get("stage"))),
                stage.get("count", 0),
                (stage.get("overall_conversion", 0) or 0) / 100,
                (stage.get("dropoff_rate", 0) or 0) / 100,
            ]
        )
    style_sheet(funnel_sheet, [28, 20, 22, 18])
    for column in ("C", "D"):
        for cell in funnel_sheet[column][1:]:
            cell.number_format = "0.0%"

    velocity_sheet = workbook.create_sheet("Hunt Velocity")
    velocity_sheet.append(
        [
            "Hunt Title",
            "Target Role",
            "Status",
            "Candidates",
            "Hired",
            "Days Open",
            "Time to Fill",
        ]
    )
    for hunt in metrics_data.get("velocity", {}).get("hunts_velocity", []):
        velocity_sheet.append(
            [
                safe(_display(hunt.get("title"))),
                safe(_display(hunt.get("target_role"))),
                safe(_display(hunt.get("status"))),
                hunt.get("total_candidates", 0),
                hunt.get("hired_count", 0),
                hunt.get("days_open", 0),
                hunt.get("time_to_fill_days", 0),
            ]
        )
    style_sheet(velocity_sheet, [32, 26, 16, 14, 12, 14, 16])

    source_sheet = workbook.create_sheet("Source Quality")
    source_sheet.append(["Source", "Candidate Count"])
    for source, count in metrics_data.get("sourcing", {}).get("channels", {}).items():
        source_sheet.append([safe(_display(source)), count])
    style_sheet(source_sheet, [30, 20])

    outreach_sheet = workbook.create_sheet("Outreach")
    outreach_sheet.append(["Category", "Metric", "Value"])
    outreach = metrics_data.get("outreach", {})
    for label, count in outreach.get("channel_counts", {}).items():
        outreach_sheet.append(["Channel", safe(_display(label)), count])
    for label, count in outreach.get("direction_counts", {}).items():
        outreach_sheet.append(["Direction", safe(_display(label)), count])
    for sequence in outreach.get("sequence_performance", []):
        sequence_name = safe(_display(sequence.get("sequence_name")))
        outreach_sheet.append(["Sequence enrolled", sequence_name, sequence.get("enrolled", 0)])
        outreach_sheet.append(["Sequence replied", sequence_name, sequence.get("replied", 0)])
        outreach_sheet.append(
            [
                "Sequence response rate",
                sequence_name,
                (sequence.get("response_rate", 0) or 0) / 100,
            ]
        )
        outreach_sheet.cell(outreach_sheet.max_row, 3).number_format = "0.0%"
    style_sheet(outreach_sheet, [24, 32, 16])

    trend_sheet = workbook.create_sheet("Trends")
    trend_sheet.append(["Date", "Candidates Sourced", "Outreach Sent", "Hires"])
    trends = metrics_data.get("trends", {})
    labels = trends.get("date_labels", [])
    sourced = trends.get("candidates_sourced", [])
    sent = trends.get("outreach_sent", [])
    hires = trends.get("hires", [])
    for index, label in enumerate(labels):
        trend_sheet.append(
            [
                safe(_display(label)),
                sourced[index] if index < len(sourced) else 0,
                sent[index] if index < len(sent) else 0,
                hires[index] if index < len(hires) else 0,
            ]
        )
    style_sheet(trend_sheet, [18, 22, 18, 12])

    for sheet in workbook.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.tabColor = teal
        if sheet.max_row > 1:
            for row_index in range(2, sheet.max_row + 1):
                if row_index % 2 == 0:
                    for cell in sheet[row_index]:
                        cell.fill = PatternFill("solid", fgColor=pale_blue)
        if sheet.title == "Executive Summary":
            for coordinate in ("A2", "A3"):
                sheet[coordinate].font = Font(color=slate, bold=True)
            for row_index in range(6, 16, 2):
                for cell in sheet[row_index][:2]:
                    cell.fill = PatternFill("solid", fgColor=pale_teal)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def generate_analytics_pdf(
    metrics_data: dict[str, Any],
    *,
    generated_at: datetime | None = None,
    scope_label: str = "All Talent Hunts",
) -> bytes:
    """Generate a compact executive analytics PDF with escaped user text."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import (
            HRFlowable,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:  # pragma: no cover - guarded by the runtime dependency
        raise RuntimeError("reportlab is required for PDF report generation") from exc

    timestamp = _generated_at(generated_at)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=42,
        title="TalentHunt OS Analytics Report",
        author="TalentHunt OS",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0f172a"),
        fontName="Helvetica-Bold",
    )
    section_style = ParagraphStyle(
        "SectionHeader",
        parent=styles["Heading2"],
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#0f172a"),
        fontName="Helvetica-Bold",
        spaceBefore=12,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#334155"),
    )
    header_style = ParagraphStyle(
        "CellHeader",
        parent=body_style,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )

    def paragraph(value: Any, style=body_style):
        return Paragraph(escape(_display(value)), style)

    def table(headers: list[str], rows: list[list[Any]], widths: list[int]):
        values = [[paragraph(label, header_style) for label in headers]]
        values.extend([[paragraph(value) for value in row] for row in rows])
        result = Table(values, colWidths=widths, repeatRows=1, hAlign="LEFT")
        result.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f1f5f9")],
                    ),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        return result

    story = [
        Paragraph("TalentHunt OS - Analytics and Intelligence", title_style),
        paragraph(f"Generated {timestamp.strftime('%Y-%m-%d %H:%M UTC')} | Scope: {scope_label}"),
        Spacer(1, 8),
        HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#0f9d94")),
    ]

    kpi = metrics_data.get("kpi", {})
    story.extend(
        [
            Paragraph("Executive KPI Overview", section_style),
            table(
                ["Metric", "Value", "Metric", "Value"],
                [
                    [
                        "Total Talent Hunts",
                        kpi.get("total_hunts", 0),
                        "Candidates Sourced",
                        kpi.get("total_sourced", 0),
                    ],
                    [
                        "Active Hunts",
                        kpi.get("active_hunts", 0),
                        "Candidates Hired",
                        kpi.get("hired_candidates", 0),
                    ],
                    [
                        "Funnel Conversion",
                        f"{kpi.get('conversion_rate', 0)}%",
                        "Average Time-to-Fill",
                        f"{kpi.get('avg_time_to_fill_days', 0)} days",
                    ],
                    [
                        "Outreach Response",
                        f"{kpi.get('response_rate', 0)}%",
                        "Recorded AI Operations",
                        kpi.get("ai_actions", 0),
                    ],
                ],
                [145, 80, 155, 80],
            ),
        ]
    )

    funnel_rows = [
        [
            stage.get("stage"),
            stage.get("count"),
            f"{stage.get('overall_conversion')}%",
            f"{stage.get('dropoff_rate')}%",
        ]
        for stage in metrics_data.get("funnel", {}).get("stages", [])
    ]
    story.extend(
        [
            Paragraph("Recruitment Funnel", section_style),
            table(
                ["Pipeline Stage", "Candidates", "Overall Conversion", "Drop-off"],
                funnel_rows,
                [150, 90, 140, 90],
            ),
        ]
    )

    velocity_rows = [
        [
            hunt.get("title"),
            hunt.get("status"),
            hunt.get("total_candidates"),
            hunt.get("hired_count"),
            hunt.get("days_open"),
            hunt.get("time_to_fill_days"),
        ]
        for hunt in metrics_data.get("velocity", {}).get("hunts_velocity", [])
    ]
    story.extend(
        [
            Paragraph("Hunt Velocity and Time-to-Fill", section_style),
            table(
                ["Hunt", "Status", "Candidates", "Hired", "Days Open", "Time to Fill"],
                velocity_rows,
                [160, 72, 70, 50, 70, 70],
            ),
        ]
    )

    ai_cost = metrics_data.get("ai_cost", {})
    story.extend(
        [
            Paragraph("AI Telemetry", section_style),
            paragraph(
                "This report includes only persisted operation and cost fields. "
                f"Recorded operations: {ai_cost.get('total_operations', 0)}; "
                f"recorded cloud cost: ${ai_cost.get('actual_cloud_cost', 0.0)}; "
                f"recorded cost saved: ${ai_cost.get('total_cost_saved', 0.0)}."
            ),
        ]
    )

    def footer(canvas, document) -> None:
        canvas.saveState()
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.setFont("Helvetica", 8)
        canvas.drawString(36, 22, "TalentHunt OS | Local confidential recruitment report")
        canvas.drawRightString(576, 22, f"Page {document.page}")
        canvas.restoreState()

    try:
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
    except Exception as exc:
        raise RuntimeError(f"PDF report generation failed: {exc}") from exc
    return buffer.getvalue()
